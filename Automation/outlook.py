import win32com.client
import pandas as pd
import os
import time
from openpyxl import load_workbook
import logging
import re
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import glob

# Global dictionary to track profile emails and their corresponding feedback
profile_email_received_times = {}



# Set up logging
logging.basicConfig(
    filename='email_checker_and_feedback.log',
    level='INFO',
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Directory to save files
SAVE_DIR = r'C:\temp'
os.makedirs(SAVE_DIR, exist_ok=True)
TABLE_FILE_PATH = os.path.join(SAVE_DIR, "candidate_table.xlsx")
FEEDBACK_FILE_PATH = os.path.join(SAVE_DIR, "profile_feedback.xlsx")
METRICS_FILE_PATH = os.path.join(SAVE_DIR, "feedback_metrics.xlsx")
LAST_RUN_TIME_FILE = os.path.join(SAVE_DIR, "feedback_metrics.txt")
# Define allowed senders for profile and feedback emails
# PROFILE_SENDERS = [
# "nagaraju.l@lyptus-tech.com",
 # "sandeep.regalla@lyptus-tech.com",
    # "profile.sender2@example.com"
# ]

# FEEDBACK_SENDERS = [
#     #   "venkatarramana@lyptus-tech.com",
#     "sandeep.regalla@lyptus-tech.com",
# "nagaraju.l@lyptus-tech.com",
# ]
#
# JOB_IDS = ["AMD_1", "INT_1"]
# client = ["intel", "amd"]
# import emailNames
with open("emailNames.txt", "r") as file:
    content = file.read()
exec(content)
FEEDBACK_SENDERS = FEEDBACK_SENDERS
PROFILE_SENDERS = PROFILE_SENDERS
JOB_IDS =JOB_IDS
client = client
REMINDER_INTERVAL = REMINDER_INTERVAL
COUNTDOWN = COUNTDOWN
REMINDER_INTERVAL = timedelta(minutes= REMINDER_INTERVAL)

def append_to_excel(df, file_path):
    logging.info(f"Appending data to {file_path}...")
    try:
        if not os.path.exists(file_path):
            df.to_excel(file_path, index=False)
            logging.info(f"File created and data written to {file_path}")
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                workbook = load_workbook(file_path)
                sheet_name = workbook.sheetnames[0]
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                start_row = len(existing_df) + 1
                df.to_excel(writer, startrow=start_row, index=False, header=False)
            logging.info(f"Data successfully appended to {file_path}")
    except Exception as e:
        logging.error(f"Failed to append data to {file_path}: {e}")

def initialize_excel(file_path, columns):
    if not os.path.exists(file_path):
        df = pd.DataFrame(columns=columns)
        df.to_excel(file_path, index=False)
        logging.info(f"Initialized new Excel file with headers at {file_path}")

def get_job_id_from_subject(subject):
    for job_id in JOB_IDS:
        if job_id in subject:
            return job_id
    return None
def get_client_name_from_subject(subject):
    for clent in client:
        clent = clent.capitalize()
        print(clent)
        print(subject)
        if clent in subject:
            print(clent)
            return clent



def get_first_five_lines(body):
    lines = body.splitlines()
    first_five_lines = [line.strip() for line in lines if line.strip()][:5]
    return " ".join(first_five_lines)


def calculate_metrics_and_update(last_metrics_run_time):
    logging.info("Calculating metrics and updating metrics Excel...")

    # Load the feedback data
    try:
        df_feedback = pd.read_excel(FEEDBACK_FILE_PATH)
        if df_feedback.empty:
            logging.warning("No feedback data found to calculate metrics.")
            return
    except Exception as e:
        logging.error(f"Error reading feedback file: {e}")
        return

    # Initialize metrics DataFrame
    df_metrics = pd.DataFrame(columns=["Client Name", "Job ID", "Total Profiles Received", "Selected Profiles",
                                       "Rejected Profiles", "Rejection Reasons", "Pending"])

    # Get the unique clients and job IDs
    clients = df_feedback["Client"].unique()
    job_ids = df_feedback["Job ID"].unique()

    for client in clients:
        for job_id in job_ids:
            # Filter data for the specific client and job ID
            client_data = df_feedback[(df_feedback["Client"] == client) & (df_feedback["Job ID"] == job_id)]

            if client_data.empty:
                continue

            # Calculate metrics
            total_profiles_received = len(client_data)
            selected_profiles = len(client_data[client_data['Status'] == "Selected"])
            rejected_profiles = len(client_data[client_data['Status'] == "Rejected"])
            pending_profiles = len(client_data[client_data['Status'] == "Pending"])
            print(pending_profiles)
            if pending_profiles == "" or None:
                pending_profiles = 0

            # Group rejection reasons
            rejection_reasons = client_data[client_data['Status'] == "Rejected"][
                'Rejection Reason'].value_counts().to_dict()
            rejection_reasons_str = ', '.join([f"{reason}: {count}" for reason, count in rejection_reasons.items()])

            # Prepare metrics data
            metrics_data = {
                # "Date": [datetime.now().strftime("%Y-%m-%d")],
                "Client Name": [client],
                "Job ID": [job_id],
                "Total Profiles Received": [total_profiles_received],
                "Selected Profiles": [selected_profiles],
                "Rejected Profiles": [rejected_profiles],
                "Rejection Reasons": [rejection_reasons_str],
                "Pending": [pending_profiles],
                # "Client Name": [client],
                # "Job ID": [job_id]
            }

            # Convert to DataFrame and append to the metrics DataFrame
            df_metrics = pd.concat([df_metrics, pd.DataFrame(metrics_data)], ignore_index=True)

    # Append to the metrics Excel file
    append_to_excel(df_metrics, METRICS_FILE_PATH)
    logging.info(f"Metrics data appended to {METRICS_FILE_PATH}")
# def calculate_metrics_and_update(last_metrics_run_time):
#     logging.info("Calculating metrics and updating metrics Excel...")
#
#     # Load the feedback data
#     try:
#         df_feedback = pd.read_excel(FEEDBACK_FILE_PATH)
#         if df_feedback.empty:
#             logging.warning("No feedback data found to calculate metrics.")
#             return
#     except Exception as e:
#         logging.error(f"Error reading feedback file: {e}")
#         return
#
#     # Convert the 'Timestamp' column to datetime if it exists
#     if 'Timestamp' in df_feedback.columns:
#         df_feedback['Timestamp'] = pd.to_datetime(df_feedback['Timestamp'])
#
#     # Filter data after the last metrics run time
#     # if last_metrics_run_time is not None:
#     #     df_feedback = df_feedback[df_feedback['Timestamp'] > last_metrics_run_time]
#
#     if df_feedback.empty:
#         logging.info("No new feedback data found since the last metrics run.")
#         return
#
#     # Initialize metrics DataFrame
#     df_metrics = pd.DataFrame(columns=["Date", "Total Profiles Received", "Selected Profiles",
#                                        "Rejected Profiles", "Rejection Reasons", "Pending", "Client Name", "Job ID"])
#
#     # Get the unique clients and job IDs
#     clients = df_feedback["Client"].unique()
#     job_ids = df_feedback["Job ID"].unique()
#
#     for client in clients:
#         for job_id in job_ids:
#             # Filter data for the specific client and job ID
#             client_data = df_feedback[(df_feedback["Client"] == client) & (df_feedback["Job ID"] == job_id)]
#
#             if client_data.empty:
#                 continue
#
#             # Calculate metrics
#             total_profiles_received = len(client_data)
#             selected_profiles = len(client_data[client_data['Status'] == "Selected"])
#             rejected_profiles = len(client_data[client_data['Status'] == "Rejected"])
#             pending_profiles = len(client_data[client_data['Status'] == "Pending"])
#
#             # Group rejection reasons
#             rejection_reasons = client_data[client_data['Status'] == "Rejected"][
#                 'Rejection Reason'].value_counts().to_dict()
#             rejection_reasons_str = ', '.join([f"{reason}: {count}" for reason, count in rejection_reasons.items()])
#
#             # Prepare metrics data
#             metrics_data = {
#                 "Client Name": [client],
#                 "Job ID": [job_id],
#                 "Total Profiles Received": [total_profiles_received],
#                 "Selected Profiles": [selected_profiles],
#                 "Rejected Profiles": [rejected_profiles],
#                 "Rejection Reasons": [rejection_reasons_str],
#                 "Pending": [pending_profiles],
#             }
#
#             # Convert to DataFrame and append to the metrics DataFrame
#             df_metrics = pd.concat([df_metrics, pd.DataFrame(metrics_data)], ignore_index=True)
#
#     # Append to the metrics Excel file
#     append_to_excel(df_metrics, METRICS_FILE_PATH)
#     logging.info(f"Metrics data appended to {METRICS_FILE_PATH}")
#
#     # Update the last metrics run time with the current timestamp
#     update_last_run_time()


def update_last_run_time():
    current_time = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(LAST_RUN_TIME_FILE, 'w') as f:
        f.write(current_time)
    logging.info(f"Updated last metrics run time to {current_time}")

def get_last_run_time():
    if os.path.exists(LAST_RUN_TIME_FILE):
        with open(LAST_RUN_TIME_FILE, 'r') as f:
            last_run_time_str = f.read().strip()
            return pd.Timestamp(last_run_time_str)
    else:
        return None  # No previous run time found



def send_reminder_email(client, job_id, candidateName , feedback_pending , recived_time):
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0: olMailItem
        print(type(feedback_pending))
        feedback_pending = str(feedback_pending)[0:8]
        print(type(recived_time))
        recived_time = str(recived_time)[0:19]
        mail.Subject = f"Reminder: Feedback Required for {candidateName}"
        mail.Body = (
            f"Dear Team,\n\n"
            f"This is a gentle reminder to send feedback for the profiles received for {client}, Job ID: {job_id}, CandidateName :{candidateName}.\n\n"
            f"Feedback is pending from last {feedback_pending} duration , profile mail received at {recived_time}.\n\n"
            f"Best Regards,\n"
            f"Lyptus"
        )
        attachments_dir = r'C:\temp\saved_emails'

        # Get the list of files in the directory
        files = glob.glob(os.path.join(attachments_dir, '*'))

        # Sort files by modification time, latest first
        latest_file = max(files, key=os.path.getmtime)

        # Check if the latest file exists
        if os.path.exists(latest_file):
            mail.Attachments.Add(latest_file)
            logging.info(f"Attached file {latest_file} to the reminder email.")
        else:
            logging.error(f"File {latest_file} does not exist.")
        # if attachment_path and os.path.(attachment_path):
        #     mail.Attachments.Add(attachment_path)
        #     logging.info(f"Attached file {attachment_path} to the reminder email.")

        # Send to all feedback senders
        mail.To = ";".join(FEEDBACK_SENDERS)

        # Send the email
        mail.Send()
        logging.info(f"Reminder email sent for client {client} and job ID {job_id}.")

    except Exception as e:
        logging.error(f"Failed to send reminder email: {e}")


EMAIL_SAVE_DIR = r'C:\temp\saved_emails'
os.makedirs(EMAIL_SAVE_DIR, exist_ok=True)


def save_email_as_msg(email):
    try:
        # Construct the filename for the email, ensuring it's unique and safe for saving
        subject_safe = ''.join(e for e in email.Subject if e.isalnum() or e in (' ', '_')).strip()
        save_name = f"{subject_safe}_{email.ReceivedTime.strftime('%Y%m%d_%H%M%S')}.msg"

        # Full path to save the email
        email_save_path = os.path.join(EMAIL_SAVE_DIR, save_name)

        # Save the email as a .msg file
        email.SaveAs(email_save_path, 3)  # 3 corresponds to saving as an Outlook message format (.msg)

        logging.info(f"Email saved as {email_save_path}")
    except Exception as e:
        logging.error(f"Failed to save email: {e}")
def process_email(email):
    subject1 = email['subject']
    sender = email['sender']
    print(sender, '-------------------------')
    logging.info(f"Processing email from {sender} with subject: {subject1}")

    job_id = get_job_id_from_subject(subject1)
    print(job_id)
    client = get_client_name_from_subject(subject1)
    print(type(subject1))
    subject2 =  subject1.strip("RE:")
    print(subject2)
    # RE:amd,jib_id,name
    # [re:amd , jbid , name ]
    candidateNameinfo = subject2.split(",")
    candidateName = candidateNameinfo[-1]
    print(candidateName)
    if not job_id:
        logging.warning(f"No valid job ID found in the subject: {subject1}. Skipping email.")
        return
    if not client:
        logging.warning(f"No valid Client found in the subject: {subject1}. Skipping email.")
        return
    # save_email_attachments(email)
    # candidateName = subject2.split(",")
    # print(candidateName)
    # print(sender)
    # Check if the sender is in profile or feedback senders list
    if sender in PROFILE_SENDERS and job_id in JOB_IDS:
        logging.info("Processing as Profile email")
        profile_email_received_times[(client, job_id,candidateName)] = datetime.now()
        logging.info(f"Profile email received for {client} and job ID {job_id}.")
        html_content = email["message"].HTMLBody
        if html_content:
            soup = BeautifulSoup(html_content, "html.parser")
            table = soup.find("table")
            if table:
                logging.info("Table found in the email.")
                rows = table.find_all("tr")
                table_data = []
                for row in rows:
                    cols = row.find_all("td")
                    cols = [ele.text.strip() for ele in cols]
                    table_data.append(cols)
                # df_table["Date"] = [datetime.now().strftime("%Y-%m-%d")]
                df_table = pd.DataFrame(table_data[1:], columns=table_data[1])
                df_table["Job ID"] = job_id
                # df_table["Date"] = [datetime.now().strftime("%Y-%m-%d")]
                df_table["Client"] = client
                df_table["Date"] = [datetime.now().strftime("%Y-%m-%d")]
                append_to_excel(df_table, TABLE_FILE_PATH)
                logging.info(f"Table data appended to {TABLE_FILE_PATH}")
            else:
                logging.warning("No table found in the email.")
        else:
            logging.warning("Email body is empty or not in HTML format.")

    elif sender in FEEDBACK_SENDERS and job_id in JOB_IDS:
        logging.info("Processing as Feedback email")
        body = email["message"].Body.lower()
        body = get_first_five_lines(body)
        if (client, job_id) in profile_email_received_times:
            del profile_email_received_times[(client, job_id,candidateName)]
            logging.info(f"Feedback received for {client} and job ID {job_id} and cadidatename {candidateName}.")
        # Determine the status and reason for feedback
        if "selected" in body:
            status = "Selected"
            body = body.split(".")
            rejection_reason = body[1]
            print(rejection_reason)
            logging.info("Candidate is selected")
        elif "not suitable" or "not selected" in body:
            status = "Rejected"
            body = body.split(".")
            rejection_reason = body[1]
            # rejection_reason = re.findall(r"not suitable,\s*(.*)", body, re.IGNORECASE)
            rejection_reason = body[1]
            logging.info(f"Candidate is rejected with reason: {rejection_reason}")
        else:
            status = "Pending"
            rejection_reason = "No response"
            logging.info("Candidate status is pending")

        # Prepare the feedback data
        data_feedback = {
            "Date" : [datetime.now().strftime("%Y-%m-%d")],
            "Candidate Name" : candidateName,
            "Job ID": job_id,
            "Client": client,
            "Status": status,
            "Rejection Reason" :rejection_reason,
        }
        print(client)
        print(candidateName)
        print(job_id)
        print(rejection_reason)
        print(status)

        df_feedback = pd.DataFrame([data_feedback])
        append_to_excel(df_feedback, FEEDBACK_FILE_PATH)
        logging.info(f"Feedback data appended to {FEEDBACK_FILE_PATH}")

    # Mark the email as read
    email["message"].Unread = False
    logging.info("Email marked as read")

def check_for_job_description(email):
    print("...........................................")
    body = email["message"].Body.lower()
    body1 = body
    print(body1)
    print(type(body1))
    subject1 = email['subject']
    if "job description" in body1:
        pass
    elif "jd:" in body1:
        print("pass................")
        pass
    else:
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)  # 0: olMailItem
            mail.Subject = "JD-Remainder JD missing"
            mail.Body = (
                f"Dear Team,\n\n"
                f"No job Description found please attach the JD for this profile\n\n"
                f"Best Regards,\n"
                f"Lyptus"
            )
            attachments_dir = r'C:\temp\saved_emails'

            # Get the list of files in the directory
            files = glob.glob(os.path.join(attachments_dir, '*'))

            # Sort files by modification time, latest first
            latest_file = max(files, key=os.path.getmtime)

            # Check if the latest file exists
            if os.path.exists(latest_file):
                mail.Attachments.Add(latest_file)
                logging.info(f"Attached file {latest_file} to the reminder email.")
            else:
                logging.error(f"File {latest_file} does not exist.")
            # Send to all feedback senders
            mail.To = ";".join(PROFILE_SENDERS)
            # Send the email
            mail.Send()
            logging.info(f"Reminder email sent for JD")

        except Exception as e:
            logging.error(f"Failed to send reminder email: {e}")

def check_for_reminders():
    current_time = datetime.now()
    # feedback_pending = current_time -received_time

    # Check if any profiles have passed the 10-minute threshold without feedback
    for (client, job_id , candidateName), received_time in list(profile_email_received_times.items()):
        if current_time - received_time > REMINDER_INTERVAL:
            logging.info(f"Sending reminder for client {client} and job ID {job_id}.")
            feedback_pending = current_time - received_time
            send_reminder_email(client, job_id, candidateName , feedback_pending, received_time)

            del profile_email_received_times[(client, job_id ,candidateName)]  # Remove after sending the reminder
def check_emails():
    logging.info("Checking for new emails...")

    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        accounts = outlook.Folders
        print("Available accounts:", accounts)
        for i, account in enumerate(accounts):
            print(f"{i + 1}. {account.Name}")

        # Select the second account (index 1, as it's zero-based)
        second_account = accounts.Item(3)  # Use 2 for the second account

        # Access the inbox folder of the second account
        inbox = second_account.Folders("Inbox")
        # inbox = outlook.GetDefaultFolder(6)
        # print(inbox)
        messages = inbox.Items
        print(messages)
        messages.Sort("[ReceivedTime]", True)
        message_count = 0
        for message in messages:
            if message_count >= 1:
                break

            try:
                # Use PropertyAccessor to get the SMTP address
                sender = message.Sender
                property_accessor = sender.PropertyAccessor
                smtp_address = property_accessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
                print(f"Sender's email ID: {smtp_address}")
            except Exception as e:
                print(f"Failed to retrieve email address: {e}")
            message_count += 1
        # print(messages)
        for message in messages:
            # print(message.SenderEmailAddress)

            if message.UnRead:
                save_email_as_msg(message)
                email_data = {
                    "subject": message.Subject,
                    "sender": smtp_address,
                    "message": message
                }
                check_for_job_description(email_data)# need to handle this
                process_email(email_data)
            else:
                logging.info("No more unread emails to process.")
                break

    except Exception as e:
        logging.error(f"Error checking emails: {e}")

# Initialize Excel files with correct headers
initialize_excel(FEEDBACK_FILE_PATH, ["Date","Candidate Name","Job ID","Client", "Status", "Rejection Reason"])
initialize_excel(TABLE_FILE_PATH, ["Candidate Name", "Skill", "Experience", "Current Company", "Current CTC", "Exp CTC",
                                   "Notice Period", "Current location", "Preferred location", "Any Offer in hand",
                                   "Mobile No.", "Email id", "Job ID", "Client","Date"])
initialize_excel(METRICS_FILE_PATH, [ "Client Name","Job ID" ,"Total Profiles Received", "Selected Profiles", "Rejected Profiles",
                                     "Rejection Reasons","Pending"])
import pandas as pd
import os
import win32com.client
import logging

def send_feedback_metrics_email():
    try:
        # Load the metrics data
        if os.path.exists(METRICS_FILE_PATH):
            df_metrics = pd.read_excel(METRICS_FILE_PATH)
            received_jobid = (df_metrics['Job ID'].values)
            print(received_jobid,"-----------------------------------")
            if df_metrics.empty:
                logging.error(f"No data found in {METRICS_FILE_PATH}")
                return
        else:
            logging.error(f"Metrics file {METRICS_FILE_PATH} does not exist.")
            return
            # List of known Job IDs
        known_job_ids = JOB_IDS  # Example list of Job IDs
        print(type(known_job_ids))
        # received_jobid = (df_metrics['Job_id'].values)
        # print(received_jobid)
            # Check if any job ids are missing and add them with profile count as zero
        for job_id in known_job_ids:
            if job_id not in received_jobid:
                logging.warning(f"{job_id} not found in metrics, adding with zero profiles received.")
                # Append new row with Job ID and Profiles Received = 0
                new_row = {'Client Name':'',
           'Job ID': job_id,
           'Total Profiles Received': 0,
           'Selected Profiles': 0,  # default value for pending
           'Rejected Profiles': 0,  # Add defaults for other columns if necessary
           'Rejection Reasons': 0,
            'Pending':0}

                # Print the new row dictionary
                print(new_row)

                # Convert the new row dictionary to a DataFrame
                new_row_df = pd.DataFrame([new_row])

                # Concatenate the original DataFrame with the new row DataFrame
                df_metrics = pd.concat([df_metrics, new_row_df], ignore_index=True)

                # Print the updated DataFrame to check the result
                print(df_metrics)

            # Get the latest inserted row
        # Get the latest inserted row
        latest_two_rows = df_metrics
        print(latest_two_rows)

        # Convert the latest rows to an HTML table
        html_table = latest_two_rows.to_html(index=False, border=1, justify="center")
        print(html_table)
        Date = datetime.now().strftime("%Y-%m-%d")

        # Compose the email body
        email_body = f"""
        <html>
        <body>
            <p>Dear Team,</p>
            <p>Please find below the latest profiles shortlists count report on Date : {Date}:</p>
            {html_table}
            <p>Best Regards,<br>Lyptus</p>
        </body>
        </html>
        """
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0: olMailItem

        # Get the second account
        namespace = outlook.GetNamespace("MAPI")
        accounts = namespace.Folders
        second_account = accounts.Item(2)  # Access the second account
        # Send the email
        # outlook = win32com.client.Dispatch("Outlook.Application")
        # mail = outlook.CreateItem(0)  # 0: olMailItem
        mail.Subject = "<no_reply> Feedback Metrics Report"
        mail.HTMLBody = email_body  # Use HTML body for the email
        mail.SendUsingAccount = outlook.Session.Accounts.Item(2)  # Set second account
        # Add recipient(s)
        mail.To = ";".join(FEEDBACK_SENDERS)  # Send to all feedback senders

        # Attach the metrics file
        attachment_path = METRICS_FILE_PATH
        if os.path.exists(attachment_path):
            mail.Attachments.Add(attachment_path)
            logging.info(f"Attached file {attachment_path} to the email.")
        else:
            logging.error(f"Metrics file {attachment_path} does not exist.")

        # Send the email
        mail.Send()
        logging.info("Feedback metrics email sent successfully with the latest data in the body.")

    except Exception as e:
        logging.error(f"Failed to send feedback metrics email: {e}")


# def send_feedback_metrics_email():
#     try:
#         outlook = win32com.client.Dispatch("Outlook.Application")
#         mail = outlook.CreateItem(0)  # 0: olMailItem
#         mail.Subject = "<no_reply>Feedback Metrics Report"
#         mail.Body = (
#             "Dear Team,\n\n"
#             "Please find attached the latest profiles shortlists count report.\n\n"
#             "Best Regards,\n"
#             "Lyptus"
#         )
#
#         # Add recipient(s)
#         mail.To = ";".join(FEEDBACK_SENDERS)  # Send to all feedback senders
#
#         # Attach the metrics file
#         attachment_path = METRICS_FILE_PATH
#         if os.path.exists(attachment_path):
#             mail.Attachments.Add(attachment_path)
#             logging.info(f"Attached file {attachment_path} to the email.")
#         else:
#             logging.error(f"Metrics file {attachment_path} does not exist.")
#
#         # Send the email
#         mail.Send()
#         logging.info("Feedback metrics email sent successfully.")
#
#     except Exception as e:
#         logging.error(f"Failed to send feedback metrics email: {e}")


from datetime import datetime, timedelta
profile_email_tracker = {}
REMINDER_INTERVAL = timedelta(minutes=10)  # Time after which to send a reminder if no feedback received


def main():
    # email_check_interval = 5  # in seconds
    # logging.info("Starting the email checker...")
    # while True:
    #     check_emails()
    #     time.sleep(email_check_interval)

    email_check_interval = 5  # in seconds
    logging.info("Starting the email checker...")

    # Time tracking for metrics update
    last_metrics_update_time = datetime.now()
    print(last_metrics_update_time)
    # save_attachments()
    last_metrics_run_time = get_last_run_time()
    count = 0

    while True:
        # Check emails periodically
        check_emails()

        check_for_reminders()

        # print(profile_email_received_times)

        # Check if 24 hours have passed since the last metrics update
        # if datetime.now() - last_metrics_update_time >= timedelta(seconds=300):
        #     calculate_metrics_and_update()
        #     last_metrics_update_time = datetime.now()
        #     send_feedback_metrics_email()
        #
        # # Wait for the next iteration
        if count == COUNTDOWN:
            # last_metrics_run_time = get_last_run_time()
            calculate_metrics_and_update(last_metrics_run_time)
            # last_metrics_run_time = pd.Timestamp.now()
            send_feedback_metrics_email()
            print("yes//////////////")
            count = 0
            # del FEEDBACK_FILE_PATH
            # del METRICS_FILE_PATH
            time.sleep(10)
            if os.path.exists(r"C:\temp\profile_feedback.xlsx"):
                new_file_name = datetime.now().strftime(r"C:\temp\profile_feedback_%Y-%m-%d_%H-%M.xlsx")
                os.rename(r"C:\temp\profile_feedback.xlsx", new_file_name)
            if os.path.exists(r"C:\temp\feedback_metrics.xlsx"):
                new_file_name = datetime.now().strftime(r"C:\temp\feedback_metrics_%Y-%m-%d_%H-%M-%S.xlsx")
                os.rename(r"C:\temp\feedback_metrics.xlsx" , new_file_name)
            # if os.path.exists(r"C:\temp\profile_feedback.xlsx"):
            #     new_file_name = datetime.now().strftime(r"C:\temp\profile_feedback_%Y-%m-%d_%H-%M.xlsx")
            #     os.rename(r"C:\temp\profile_feedback.xlsx", new_file_name)
        time.sleep(email_check_interval)
        # del FEEDBACK_FILE_PATH
        # del METRICS_FILE_PATH
        count = count+5
if __name__ == "__main__":
    main()