
import win32com.client
import pandas as pd
import os
import logging
import re
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime, timedelta


profile_email_received_times = {}

# Set up logging
logging.basicConfig(
    filename='email_checker_and_feedback.log',
    level='INFO',
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Directory to save files
SAVE_DIR = r'C:\temp'
os.makedirs(SAVE_DIR, exist_ok=True)
TABLE_FILE_PATH = os.path.join(SAVE_DIR, "candidate_table.xlsx")

# Load global variables from external file
with open("attribute_names.txt", "r") as file:
    content = file.read()
exec(content)

# Initialize variables loaded from attribute_names.txt
FEEDBACK_SENDERS = FEEDBACK_SENDERS
PROFILE_SENDERS = PROFILE_SENDERS
JOB_IDS = JOB_IDS
client = client
REMINDER_INTERVAL_MINUTES = REMINDER_INTERVAL  # Avoiding overwriting global REMINDER_INTERVAL
REMINDER_INTERVAL = timedelta(minutes=REMINDER_INTERVAL_MINUTES)

# Function to append data to Excel
def append_to_excel(df, file_path):
    logging.info(f"Appending data to {file_path}...")
    try:
        if not os.path.exists(file_path):
            df.to_excel(file_path, index=False)
            logging.info(f"File created and data written to {file_path}")
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                workbook = load_workbook(file_path)
                sheet_name = workbook.sheetnames[0]
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                start_row = len(existing_df) + 1
                df.to_excel(writer, startrow=start_row, index=False, header=False)
            logging.info(f"Data successfully appended to {file_path}")
    except Exception as e:
        logging.error(f"Failed to append data to {file_path}: {e}")


# Function to initialize Excel file with column headers
def initialize_excel(file_path, columns):
    if not os.path.exists(file_path):
        df = pd.DataFrame(columns=columns)
        df.to_excel(file_path, index=False)
        logging.info(f"Initialized new Excel file with headers at {file_path}")

# Function to extract job_id from email subject
def get_job_id_from_subject(subject):
    for job_id in JOB_IDS:
        if job_id in subject:
            return job_id
    return None

# Function to extract client name from email subject
def get_client_name_from_subject(subject):
    for client_name in client:
        formatted_client = client_name.capitalize()
        if formatted_client in subject:
            return formatted_client
    return None


def get_first_five_lines(body):
    lines = body.splitlines()
    first_five_lines = [line.strip() for line in lines if line.strip()][:5]
    return " ".join(first_five_lines)

def process_email(email):
    subject1 = email['subject']
    sender = email['sender']
    print(sender, '-------------------------')
    logging.info(f"Processing email from {sender} with subject: {subject1}")

    job_id = get_job_id_from_subject(subject1)
    print(job_id)
    client = get_client_name_from_subject(subject1)
    print(type(subject1))
    subject2 =  subject1.strip("RE:")
    print(subject2)
    candidateNameinfo = subject2.split(",")
    candidateName = candidateNameinfo[-1]
    print(candidateName)
    if not job_id:
        logging.warning(f"No valid job ID found in the subject: {subject1}. Skipping email.")
        return
    if not client:
        logging.warning(f"No valid Client found in the subject: {subject1}. Skipping email.")
        return

    if sender in PROFILE_SENDERS and job_id in JOB_IDS:
        logging.info("Processing as Profile email")
        profile_email_received_times[(client, job_id,candidateName)] = datetime.now()
        logging.info(f"Profile email received for {client} and job ID {job_id}.")
        html_content = email["message"].HTMLBody
        if html_content:
            soup = BeautifulSoup(html_content, "html.parser")
            table = soup.find("table")
            if table:
                logging.info("Table found in the email.")
                rows = table.find_all("tr")
                table_data = []
                for row in rows:
                    cols = row.find_all("td")
                    cols = [ele.text.strip() for ele in cols]
                    table_data.append(cols)
                # df_table["Date"] = [datetime.now().strftime("%Y-%m-%d")]
                df_table = pd.DataFrame(table_data[1:], columns=table_data[1])
                df_table["Job ID"] = job_id
                # df_table["Date"] = [datetime.now().strftime("%Y-%m-%d")]
                df_table["Client"] = client
                df_table["Date"] = [datetime.now().strftime("%Y-%m-%d")]
                append_to_excel(df_table, TABLE_FILE_PATH)
                logging.info(f"Table data appended to {TABLE_FILE_PATH}")
            else:
                logging.warning("No table found in the email.")
        else:
            logging.warning("Email body is empty or not in HTML format.")


    # Mark the email as read
    email["message"].Unread = False
    logging.info("Email marked as read")
def check_emails():
    logging.info("Checking for new emails...")

    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        accounts = outlook.Folders
        print("Available accounts:", accounts)
        for i, account in enumerate(accounts):
            print(f"{i + 1}. {account.Name}")

        # Select the second account (index 1, as it's zero-based)
        second_account = accounts.Item(1)  # Use 2 for the second account

        # Access the inbox folder of the second account
        inbox = second_account.Folders("Inbox")
        # inbox = outlook.GetDefaultFolder(6)
        # print(inbox)
        messages = inbox.Items
        print(messages)
        messages.Sort("[ReceivedTime]", True)
        message_count = 0
        for message in messages:
            if message_count >= 1:
                break

            try:
                # Use PropertyAccessor to get the SMTP address
                sender = message.Sender
                property_accessor = sender.PropertyAccessor
                smtp_address = property_accessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
                print(f"Sender's email ID: {smtp_address}")
            except Exception as e:
                print(f"Failed to retrieve email address: {e}")
            message_count += 1
        # print(messages)
        for message in messages:
            # print(message.SenderEmailAddress)

            if message.UnRead:
                # save_email_as_msg(message)
                email_data = {
                    "subject": message.Subject,
                    "sender": smtp_address,
                    "message": message
                }
                # check_for_job_description(email_data)# need to handle this
                process_email(email_data)
            else:
                logging.info("No more unread emails to process.")
                break

    except Exception as e:
        logging.error(f"Error checking emails: {e}")


initialize_excel(TABLE_FILE_PATH, ["Candidate Name", "Skill", "Experience", "Current Company", "Current CTC", "Exp CTC",
                                   "Notice Period", "Current location", "Preferred location", "Any Offer in hand",
                                   "Mobile No.", "Email id", "Job ID", "Client","Date"])


import time  # This is necessary for the time.sleep() function
import logging
from datetime import datetime, timedelta
def main():

    email_check_interval = 5  # in seconds
    logging.info("Starting the email checker...")


    last_metrics_update_time = datetime.now()
    print(last_metrics_update_time)

    count = 0

    while True:
        # Check emails periodically
        check_emails()

        if count == COUNTDOWN:

            print("yes//////////////")
            count = 0

            time.sleep(10)

        time.sleep(email_check_interval)

        count = count+5
if __name__ == "__main__":
    main()